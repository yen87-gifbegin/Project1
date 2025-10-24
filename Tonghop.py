#Tổng hợp file excel thành report.json
import openpyxl, json

workbook= openpyxl.load_workbook("Bai6_login.xlsx")
sheet=workbook["student"]
students=[] #khai báo list rỗng, tạo list rỗng để lưu file json
student_2 =[]
for row in sheet.iter_rows(min_row= 2, values_only=True):
    stt,ID, Tên , Điểm = row
#Cần dùng Try..except cho nh lỗi: ko tồn tại file hoặc sheet, tên file tên sheet sai
    try:
        Score = float(Điểm) # đổi kiểu
        No= int(stt)
    except Exception as e:
        print(e)
    if Score >=9::
        students.append({
            "ID" : ID,
            "Name" : Tên,
            "Level" :"Xuất sắc"
        }) 
    elif Score >=7:
        students.append ({
            "ID" : ID,
            "Name" : Tên,
            "Level" :"Khá"   
        })
    elif Score >=5:
        students.append({
            "ID" : ID,
            "Name" : Tên,
            "Level" :"Trung bình" 
        })
    else: 
        students.append({
            "ID" : ID,
            "Name" : Tên,
            "Level" :"Yếu"     
        })
#print (students)
#elif Score
with open("report.json","w",encoding="utf-8") as f:
    json.dump(students,f,ensure_ascii= False, indent=4 )#Tạo file json để lưu lại (chuyển đổi ) nếu ensure_ascii = True thì mọi ký tự tiếng việt sẽ bị incake
print("Tạo file Json thành công")   
# for row2 in sheet.iter_rows(min_row= 2, values_only= True):
#     stt,ID, Tên , Điểm = row2
#     try:
#         Score = float(Điểm) # đổi kiểu
#         No= int(stt)
#     except Exception as e:
#         print(e)
#     if 7 <= Score < 9:
#         student_2.append ({
#              "ID" : ID,
#             "Name" : Tên,
#             "Level" :"Khá"
#         })

# with open("report2.json","w",encoding="utf-8") as f:
#     json.dump(student_2,f) #,ensure_ascii= False, indent=4 )
